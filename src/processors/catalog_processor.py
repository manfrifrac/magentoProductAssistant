import pandas as pd
from typing import Union, Dict, List, Any
import logging
from config import Config
from src.context.product_context import ProductContext
from .size_attribute_processor import SizeAttributeProcessor
from pathlib import Path
import re
import openpyxl

class CatalogProcessor:
    def __init__(self, config: Config):
        self.config = config
        self.context_mapping_file = config.context_mapping_file
        self.magento_mapping_file = config.mapping_file  # Add this line to fix the error
        self.field_mappings = []
        self.magento_mappings: Dict[str, Dict[str, List[str]]] = {}
        self.required_fields: List[str] = []
        self.default_values: Dict[str, str] = {}
        self.image_fields = {
            'base_image': 'first',
            'small_image': 'first', 
            'thumbnail_image': 'first',
            'additional_images': 'all'
        }
        self.size_mapping = {
            'default': ['size', 'product_size'],
            'apparel': ['clothing_size', 'garment_size'],
            'shoes': ['shoe_size', 'footwear_size'],
            'accessories': ['accessory_size']
        }
        
        self.standard_sizes = {
            'apparel': ['XS', 'S', 'M', 'L', 'XL', '2XL', '3XL'],
            'shoes': [str(x) for x in range(35, 47)],  # European sizes
            'accessories': ['ONE SIZE']
        }

        # Supplier-specific size field mappings
        self.supplier_size_fields = {
            'guirca': ['taglia', 'size'],
            'widmann': ['size', 'dimension']
            # Add other suppliers as needed
        }
        
        # Magento attribute set mappings based on size patterns
        self.attribute_set_rules = {
            'apparel': {
                'patterns': ['XS', 'S', 'M', 'L', 'XL', '2XL'],
                'attribute_set': 'clothing_sizes'
            },
            'numeric': {
                'patterns': [str(x) for x in range(30, 50)],  # For shoe sizes
                'attribute_set': 'shoe_sizes'
            },
            'kids': {
                'patterns': ['2-4Y', '4-6Y', '6-8Y', '8-10Y'],
                'attribute_set': 'kids_sizes'
            }
        }
        
        # Initialize SizeAttributeProcessor with config path
        config_path = Path(__file__).parent.parent.parent / 'config' / 'supplier_size_mapping.json'
        self.size_processor = SizeAttributeProcessor(config_path)

    def _strip_image_url(self, image_path: str) -> str:
        """Strip URL from image path to get just the filename"""
        if pd.isna(image_path) or not image_path:
            return ''
        # Split on both forward and back slashes and take the last part
        return image_path.replace('\\', '/').split('/')[-1]

    def _extract_hyperlink(self, cell_value: Any) -> str:
        """Extract hyperlink from Excel cell or value"""
        try:
            # Handle openpyxl Cell object
            if hasattr(cell_value, 'hyperlink') and cell_value.hyperlink:
                return cell_value.hyperlink.target
                
            # Handle tuple format (text, url)
            if isinstance(cell_value, tuple) and len(cell_value) == 2:
                return cell_value[1]
                
            # Handle string with URL
            if isinstance(cell_value, str):
                # Try to find URLs in text
                urls = re.findall(r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', cell_value)
                if urls:
                    return urls[0]
                    
                # Check if it's a HYPERLINK formula
                if '=HYPERLINK' in cell_value:
                    match = re.search(r'=HYPERLINK\("([^"]+)"', cell_value)
                    if match:
                        return match.group(1)
                        
            return ''
            
        except Exception as e:
            logging.warning(f"Error extracting hyperlink: {e}")
            return ''

    def load_mapping(self) -> None:
        """Load both context and Magento field mappings"""
        try:
            # Load context mappings
            context_df = pd.read_csv(self.context_mapping_file)
            self.field_mappings = context_df.to_dict('records')
            logging.info(f"Loaded {len(self.field_mappings)} context mappings")
            
            # Load Magento mappings
            magento_df = pd.read_csv(self.magento_mapping_file)
            
            # Get required fields and default values
            suppliers = [col for col in magento_df.columns if col.lower() in ['guirca', 'widmann','espa']]
            
            for supplier in suppliers:
                supplier_key = supplier.lower()
                self.magento_mappings[supplier_key] = {}
                
                for _, row in magento_df.iterrows():
                    magento_field = row['magento_field']
                    supplier_field = row[supplier]
                    
                    if pd.notna(supplier_field):
                        if magento_field not in self.magento_mappings[supplier_key]:
                            self.magento_mappings[supplier_key][magento_field] = []
                        self.magento_mappings[supplier_key][magento_field].append(supplier_field)
            
            logging.info(f"Loaded Magento mappings for suppliers: {list(self.magento_mappings.keys())}")
            
        except Exception as e:
            logging.error(f"Error loading mappings: {e}")
            raise

    def determine_attribute_set(self, size_value: str) -> str:
        """Determine which Magento attribute set to use based on size value"""
        if not size_value:
            return 'default'
            
        size_value = str(size_value).upper().strip()
        
        for set_type, rules in self.attribute_set_rules.items():
            if any(pattern in size_value for pattern in rules['patterns']):
                return rules['attribute_set']
                
        return 'default'

    def process_size_attribute(self, row: Dict, magento_row: Dict, supplier_name: str) -> None:
        """Process size attribute using the size processor"""
        # Get mapped size value from magento_row
        logging.debug(f"Processing size attribute for row: {row}")
        if 'size' in magento_row:
            size_info = self.size_processor.process_size(
                product_data=magento_row,
                supplier=supplier_name
            )
            magento_row.update(size_info)
            logging.debug(f"Updated magento_row with size info: {magento_row}")
        else:
            # Set defaults if no size found
            logging.warning(f"No size field found in magento_row: {magento_row}")
            magento_row.update({
                'size': '',
                'size_set': 'default',
                'size_type': 'default'
            })

    def standardize_size(self, size_value: str, store_type: str) -> str:
        """Standardize size values according to store type"""
        if store_type in self.standard_sizes:
            if size_value in self.standard_sizes[store_type]:
                return size_value
        return size_value

    def _process_images(self, row: Dict, magento_row: Dict, mapping: Dict, supplier_key: str) -> None:
        """Process all image fields from supplier data"""
        # Process each image field according to mapping
        for magento_field, handling in self.image_fields.items():
            supplier_fields = mapping.get(magento_field, [])  # Get supplier fields from mapping
            
            if magento_field == 'additional_images':
                # Combine all additional images into comma-separated string
                images = []
                for field in supplier_fields:
                    if field in row and pd.notna(row[field]):
                        image = self._strip_image_url(str(row[field]))
                        if image:
                            images.append(image)
                magento_row[magento_field] = ','.join(images)
            else:
                # For single image fields, take first valid image
                for field in supplier_fields:
                    if field in row and pd.notna(row[field]):
                        magento_row[magento_field] = self._strip_image_url(str(row[field]))
                        break

    def process_catalog(self, supplier_name: str, file_path: str) -> Union[pd.DataFrame, None]:
        try:
            # Load workbook with openpyxl to preserve hyperlinks
            wb = openpyxl.load_workbook(file_path, data_only=False)  # Keep formulas
            ws = wb.active
            
            # Convert worksheet to DataFrame while preserving hyperlinks
            data = []
            headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
            
            # Process data rows
            for row in ws.iter_rows(min_row=2):
                row_data = {}
                for idx, (header, cell) in enumerate(zip(headers, row)):
                    if header:  # Only process cells with valid headers
                        value = cell.value
                        # Handle hyperlinks
                        if cell.hyperlink:
                            if header.lower() in ['link', 'image']:
                                value = cell.hyperlink.target
                        row_data[header] = value if value is not None else ''
                if row_data:  # Only append if we have data
                    data.append(row_data)

            # Create DataFrame
            df = pd.DataFrame(data)
            logging.info(f"Created DataFrame with {len(df)} rows and columns: {df.columns.tolist()}")
            
            supplier_key = supplier_name.lower()
            if supplier_key not in self.magento_mappings:
                logging.error(f"No Magento mapping found for supplier {supplier_name}")
                return None

            # Rest of the processing remains the same...
            # ...existing code...

            # Create DataFrame and ensure all values are strings
            df = pd.DataFrame(data)
            df = df.fillna('')  # Replace NaN with empty string
            
            # Convert all column names to strings and strip whitespace
            df.columns = df.columns.astype(str).str.strip()
            
            # Log found hyperlinks for debugging
            image_cols = [col for col in df.columns if any(x in col.lower() for x in ['link', 'image'])]
            for col in image_cols:
                links = df[col][df[col].astype(str).str.contains('http', na=False)].tolist()
                logging.info(f"Found {len(links)} hyperlinks in {col} column")
            
            # Process each row
            supplier_key = supplier_name.lower()
            
            if supplier_key not in self.magento_mappings:
                logging.error(f"No Magento mapping found for supplier {supplier_name}")
                return None

            mapping = self.magento_mappings[supplier_key]
            result_data = []
            
            # Process each row as a dictionary
            for idx, row_series in df.iterrows():
                try:
                    row = row_series.to_dict()  # Convert Series to dict
                    magento_row = {'supplier': supplier_name}
                    
                    # Create ProductContext instance for this row
                    product_context = ProductContext(
                        product_data=row,
                        supplier=supplier_name,
                        mapping_file=self.context_mapping_file
                    )

                    # Get context information as a single formatted string
                    context_data = product_context.get_context()
                    magento_row['product_context'] = context_data
                    
                    # First, try to get the SKU
                    sku_fields = mapping.get('sku', [])
                    sku = None
                    for field in sku_fields:
                        if field in row and pd.notna(row[field]):
                            sku = str(row[field]).strip()
                            magento_row['sku'] = sku
                            break
                    
                    if not sku:
                        logging.warning(f"Skipping row {idx}: No SKU found")
                        continue
                    
                    # Process regular fields
                    for magento_field, supplier_fields in mapping.items():
                        if magento_field not in self.image_fields and magento_field != 'sku':
                            for field in supplier_fields:
                                if field in row and pd.notna(row[field]):
                                    # Special handling for ESPA size and color
                                    if supplier_key == 'espa' and field == 'Color-Size':
                                        color_size = str(row[field]).strip()
                                        # Extract size from Color-Size field (format: "size-color")
                                        if '-' in color_size:
                                            size_value = color_size.split('-')[0].strip()
                                            color_value = color_size.split('-')[1].strip()
                                            if magento_field == 'size':
                                                magento_row[magento_field] = size_value
                                            elif magento_field == 'color':
                                                magento_row[magento_field] = color_value
                                        continue
                                    magento_row[magento_field] = str(row[field]).strip()
                                    break
                    
                    # Special handling for ESPA categories based on Division field
                    if supplier_key == 'espa':
                        division = str(row.get('Division', '')).strip().lower()
                        
                        # Map ESPA divisions to Magento categories
                        category_mapping = {
                            'accessories': 'accessories',
                            'basic': 'basic_wear',
                            'beachwear': 'beachwear',
                            'sportswear': 'sportswear',
                            'underwear': 'underwear'
                        }
                        
                        # Set category based on division
                        magento_row['categories'] = category_mapping.get(division, 'default')
                        
                        # Process Color-Size field
                        color_size_field = next((f for f in row.keys() if 'Color-Size' in f), None)
                        if color_size_field and pd.notna(row[color_size_field]):
                            color_size = str(row[color_size_field]).strip()
                            if '-' in color_size:
                                size_value = color_size.split('-')[0].strip()
                                color_value = color_size.split('-')[1].strip()
                                magento_row['size'] = size_value
                                magento_row['color'] = color_value
                                
                                # Determine size type based on size value
                                size_info = self.size_processor.process_size(
                                    product_data={'size': size_value},
                                    supplier=supplier_name
                                )
                                magento_row.update(size_info)
                    else:
                        # Original size processing for other suppliers
                        if 'size' in magento_row:
                            size_info = self.size_processor.process_size(
                                product_data=magento_row,
                                supplier=supplier_name,
                                original_row=row
                            )
                            magento_row.update(size_info)
                    
                    # Ensure size fields are present
                    if 'size' not in magento_row:
                        magento_row.update({
                            'size': '',
                            'size_set': 'default',
                            'size_type': 'default'
                        })
                    
                    # Process images
                    self._process_images(row, magento_row, mapping, supplier_key)
                    
                    result_data.append(magento_row)
                    
                except Exception as row_error:
                    logging.error(f"Error processing row {idx}: {row_error}")
                    continue
                    
            if not result_data:
                logging.warning(f"No valid data processed for supplier {supplier_name}")
                return None
                
            # Convert to DataFrame and verify size columns
            result_df = pd.DataFrame(result_data)
            required_columns = ['size', 'size_set', 'size_type']
            missing_columns = [col for col in required_columns if col not in result_df.columns]
            
            if missing_columns:
                logging.error(f"Missing required columns in output: {missing_columns}")
                return None
                
            # Verify we have size data
            empty_sizes = result_df['size'].isna().sum()
            if empty_sizes > 0:
                logging.warning(f"Found {empty_sizes} rows with missing size information")
                
            logging.info(f"Processed {len(result_df)} rows with size data for supplier {supplier_name}")
            return result_df
            
        except Exception as e:
            logging.error(f"Error processing catalog for supplier {supplier_name}: {e}")
            return None

    def process_all_catalogs(self) -> None:
        """Process all catalogs and create a single Magento-compatible output file"""
        try:
            input_dir = self.config.input_folder
            all_data = []
            
            supplier_dirs = [d for d in input_dir.glob("*") if d.is_dir()]
            
            if not supplier_dirs:
                logging.warning(f"No supplier directories found in {input_dir}")
                return
                
            for supplier_dir in supplier_dirs:
                supplier_name = supplier_dir.name
                excel_files = list(supplier_dir.glob("*.xlsx"))
                
                if not excel_files:
                    continue
                    
                for file_path in excel_files:
                    result_df = self.process_catalog(supplier_name, str(file_path))
                    if result_df is not None:
                        all_data.append(result_df)
            
            if all_data:
                # Combine all data
                final_df = pd.concat(all_data, ignore_index=True)
                
                # Verify size columns are present
                size_columns = ['size', 'size_set', 'size_type']
                for col in size_columns:
                    if col not in final_df.columns:
                        logging.error(f"Missing {col} column in final output")
                        return
                
                # Remove duplicates based on SKU
                final_df = final_df.drop_duplicates(subset=['sku'], keep='first')
                
                # Log size data statistics
                logging.info(f"Size distribution in final output:\n{final_df['size_set'].value_counts()}")
                
                # Save to single output file
                final_df.to_csv(self.config.output_file, index=False)
                logging.info(f"Saved combined catalog to {self.config.output_file} with {len(final_df)} rows")
                
        except Exception as e:
            logging.error(f"Error in process_all_catalogs: {e}")
            raise

    def process_row(self, row):
        # ...existing code...
        
        # Extract size information for ESPA
        if row.get('supplier', '').lower() == 'espa':
            # Use size-color field as source for size
            size = self.size_processor.extract_size(
                row.get('sku', ''), 
                row.get('size-color', ''),
                row.get('supplier', '')
            )
            
            size_set, size_type = self.size_processor.get_size_set_and_type(
                row.get('description', '')
            )
            
            row['size'] = size if size else ''
            row['size_set'] = size_set
            row['size_type'] = size_type
                
        return row
